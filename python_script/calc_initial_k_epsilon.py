"""
calc_initial_k_epsilon.py

Calculate initial turbulent kinetic energy (k) and dissipation rate (epsilon)
for OpenFOAM initial/boundary conditions.

Two regions:
  1. Ambient air (domain internalField): U=0.1 m/s, I=1%
  2. Nozzle inlet (BC): U_inject from all experimental flow rates, I=5%

Formulas (Versteeg & Malalasekera 2007):
  k       = 1.5 * (U * I)^2
  epsilon = C_mu^(3/4) * k^(3/2) / l
  C_mu    = 0.09

Length scales:
  Nozzle  : l = 0.07 * D  (pipe/nozzle fully-developed turbulence)
  Ambient : l = L_AMBIENT  (representative room-scale turbulence, user-set)

Nozzle velocity correction (ideal-gas, isobaric):
  Q_inject = Q_N * (T_inject / T_normal)   [T_normal = 273.15 K, NL/min]
  U_inject = Q_inject / A_nozzle
"""

import math
import os

# ── Constants ────────────────────────────────────────────────────────────────
C_MU        = 0.09
T_normal    = 273.15    # K  (0 °C, defines NL/min)
T_inject    = 286.15    # K  (13 °C, actual injection temperature)
T_ratio     = T_inject / T_normal

# ── User-adjustable parameters ───────────────────────────────────────────────
U_AMBIENT   = 0.1       # m/s  — ambient air velocity (initial internal field)
I_AMBIENT   = 0.01      # —    — ambient turbulence intensity (1%)
L_AMBIENT   = 0.1       # m    — ambient characteristic length scale

I_NOZZLE    = 0.05      # —    — nozzle turbulence intensity (5%)

# ── Formula helpers ──────────────────────────────────────────────────────────
def calc_k(U, I):
    return 1.5 * (U * I) ** 2

def calc_epsilon(k, l):
    return C_MU ** 0.75 * k ** 1.5 / l


# ═══════════════════════════════════════════════════════════════════════════
# 1. Ambient air
# ═══════════════════════════════════════════════════════════════════════════
k_amb  = calc_k(U_AMBIENT, I_AMBIENT)
ep_amb = calc_epsilon(k_amb, L_AMBIENT)

print("=" * 65)
print("1. Ambient air — internalField")
print("=" * 65)
print(f"   U          = {U_AMBIENT:.3f} m/s")
print(f"   I          = {I_AMBIENT*100:.1f} %")
print(f"   l          = {L_AMBIENT:.4f} m  (representative room-scale length)")
print(f"   k          = {k_amb:.4e} m²/s²")
print(f"   epsilon    = {ep_amb:.4e} m²/s³")
print(f"\n   OpenFOAM entries:")
print(f"     internalField   uniform {k_amb:.4e};   // k")
print(f"     internalField   uniform {ep_amb:.4e};   // epsilon")
print()


# ═══════════════════════════════════════════════════════════════════════════
# 2. Nozzle inlet — same case matrix as calc_nozzle_exit_velocity.py
# ═══════════════════════════════════════════════════════════════════════════
cases = {
    1: [218.1, 104.0, 62.4, 20.8, 10.4],   # NL/min, 1 m³ enclosure
    2: [218.1,  73.0, 20.8,  5.2],          # NL/min, 2 m³ enclosure
}
diameters_mm = [4.0, 27]
areas = {D: math.pi * (D * 1e-3) ** 2 / 4 for D in diameters_mm}

print("=" * 65)
print("2. Nozzle inlet — boundary condition (turbulentMixingLengthFrequencyInlet")
print("   or fixed k/epsilon)")
print("=" * 65)
print(f"   I          = {I_NOZZLE*100:.0f} %")
print(f"   l          = 0.07 × D  (pipe turbulence, V&M 2007)")
print(f"   T_inject   = {T_inject} K  |  T_ratio = {T_ratio:.4f}")
print()

col_w = [6, 8, 8, 12, 12, 12, 12]
header = (f"{'Vol':>6}  {'D[mm]':>8}  {'Q[NL/m]':>8}"
          f"  {'U_inj[m/s]':>10}  {'l[m]':>8}"
          f"  {'k[m²/s²]':>12}  {'eps[m²/s³]':>12}")
print(header)
print("-" * len(header))

for vol, Q_list in cases.items():
    for D in diameters_mm:
        A = areas[D]
        l = 0.07 * D * 1e-3
        for Q in Q_list:
            Q_m3s   = Q / 60.0 / 1000.0
            U_inj   = Q_m3s / A * T_ratio
            k_n     = calc_k(U_inj, I_NOZZLE)
            ep_n    = calc_epsilon(k_n, l)
            print(f"{vol:>6}  {D:>8.1f}  {Q:>8.1f}"
                  f"  {U_inj:>10.3f}  {l:>8.5f}"
                  f"  {k_n:>12.4e}  {ep_n:>12.4e}")
    print()

# ── Save Excel ────────────────────────────────────────────────────────────────
import pandas as pd

script_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path  = os.path.join(script_dir, "initial_k_epsilon.xlsx")

# Sheet 1: ambient
df_ambient = pd.DataFrame([{
    "Region":                "ambient air",
    "U [m/s]":               U_AMBIENT,
    "I [-]":                 I_AMBIENT,
    "l [m]":                 L_AMBIENT,
    "k [m²/s²]":             k_amb,
    "epsilon [m²/s³]":       ep_amb,
    "OF internalField k":    f"uniform {k_amb:.4e}",
    "OF internalField eps":  f"uniform {ep_amb:.4e}",
}])

# Sheet 2: nozzle — one row per (enclosure, diameter, flow rate)
nozzle_rows = []
for vol, Q_list in cases.items():
    for D in diameters_mm:
        A = areas[D]
        l = 0.07 * D * 1e-3
        for Q in Q_list:
            Q_m3s = Q / 60.0 / 1000.0
            U_inj = Q_m3s / A * T_ratio
            k_n   = calc_k(U_inj, I_NOZZLE)
            ep_n  = calc_epsilon(k_n, l)
            nozzle_rows.append({
                "Enclosure [m³]":   vol,
                "D [mm]":           D,
                "Q_N [NL/min]":     Q,
                "U_inject [m/s]":   round(U_inj, 4),
                "I [-]":            I_NOZZLE,
                "l [m]":            round(l, 6),
                "k [m²/s²]":        k_n,
                "epsilon [m²/s³]":  ep_n,
            })
df_nozzle = pd.DataFrame(nozzle_rows)

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    df_ambient.to_excel(writer, index=False, sheet_name="ambient")
    df_nozzle.to_excel(writer,  index=False, sheet_name="nozzle")

    # ── Format: scientific notation for k and epsilon columns ────────────────
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    sci_fmt = '0.0000E+00'

    for sheet_name, df in [("ambient", df_ambient), ("nozzle", df_nozzle)]:
        ws = writer.sheets[sheet_name]

        # Auto-width all columns
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(col_name)),
                         max((len(str(v)) for v in df[col_name]), default=0))
            ws.column_dimensions[col_letter].width = max_len + 4

        # Header style: bold, light-blue fill
        header_fill = PatternFill("solid", fgColor="BDD7EE")
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Scientific notation for k and epsilon columns
        for col_idx, col_name in enumerate(df.columns, start=1):
            if "k [" in col_name or "epsilon" in col_name:
                col_letter = get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=2,
                                        min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = sci_fmt

print(f"Excel saved: {xlsx_path}")
